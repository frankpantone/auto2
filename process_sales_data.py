import pandas as pd
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from collections import Counter
import numpy as np
# Use xlsxwriter for charting capabilities
import xlsxwriter

# Read the CSV file
df = pd.read_csv('raw.csv')

# Pre-process specific date columns for better Excel compatibility
date_columns = ['Created Date', 'Accepted by Carrier Date', 'Carrier Pickup Scheduled At', 'Delivery Completed At', 'Adjusted delivery date']
for col in date_columns:
    if col in df.columns:
        # Convert string dates to datetime objects before writing to Excel
        df[col] = pd.to_datetime(df[col], errors='coerce', format='%m/%d/%y')
        # Format back to string for display but now Excel will recognize as dates
        df[col] = df[col].dt.strftime('%m/%d/%Y')

# Print sample data to debug
print("Sample Vehicle delivery date values:")
sample_dates = df['Vehicle delivery date'].dropna().sample(10).tolist()
for date in sample_dates:
    print(f"  {date} (type: {type(date)})")

# Check if Vehicle delivery date is missing and populate with Adjusted delivery date where needed
df['Vehicle delivery date'] = df.apply(
    lambda row: row['Adjusted delivery date'] if pd.isna(row['Vehicle delivery date']) or str(row['Vehicle delivery date']).strip() == '' else row['Vehicle delivery date'],
    axis=1
)

# Standardize date format - Fix the date formatting issues
def standardize_date(date_val):
    if pd.isna(date_val) or str(date_val).strip() == '':
        return 'Unknown'
    
    date_str = str(date_val).strip()
    
    # If it's already in a date format with slashes, parse and reformat consistently
    if '/' in date_str:
        parts = date_str.split('/')
        if len(parts) == 3:
            try:
                month = int(parts[0])
                day = int(parts[1])
                year = parts[2]
                # Handle 2-digit years and convert to 4-digit
                if len(year) == 2:
                    year = '20' + year  # Assuming all dates are in the 2000s
                # Always format as MM/DD/YYYY with consistent leading zeros
                return f"{month:02d}/{day:02d}/{year}"
            except ValueError:
                # If conversion fails, return as is
                pass
    
    # If it doesn't match expected format, just return as is
    return date_str

# Apply standardization
df['Vehicle delivery date'] = df['Vehicle delivery date'].apply(standardize_date)

print("\nAfter standardization - unique delivery dates:")
all_dates = sorted(df['Vehicle delivery date'].unique())
print(f"Total unique dates: {len(all_dates)}")
print("First 5 dates:", all_dates[:5])
print("Last 5 dates:", all_dates[-5:])

# Ensure numeric columns are properly converted before writing to Excel
# This helps with proper formatting later
columns_to_convert = {
    'Carrier Price Per Vehicle': 'float',
    'Tariff Per Vehicle': 'float',
    'Total Carrier Price': 'float',
    'Customer Payment Total Tariff': 'float',
    'Pickup ZIP': 'int',  # Ensure ZIP codes are numeric
    'Delivery ZIP': 'int'  # Ensure ZIP codes are numeric
}

for col, dtype in columns_to_convert.items():
    if col in df.columns:
        # For numeric columns, ensure they're properly converted
        if dtype == 'float':
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        elif dtype == 'int':
            # For ZIP codes, first convert to numeric and then to integer
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)

# For columns containing currency with $ symbol, preprocess them
if 'Total Price per Mile' in df.columns:
    df['Total Price per Mile'] = df['Total Price per Mile'].apply(
        lambda x: float(str(x).replace('$', '').replace(',', '')) if isinstance(x, str) and '$' in x else x
    )
    df['Total Price per Mile'] = pd.to_numeric(df['Total Price per Mile'], errors='coerce').fillna(0)

# Create Excel workbook using xlsxwriter for charting capabilities
workbook = xlsxwriter.Workbook('sales_report.xlsx', {'nan_inf_to_errors': True})

# Create the main data worksheet
worksheet = workbook.add_worksheet('Sales Data')

# Write the dataframe to the worksheet
# First write headers
for col, header in enumerate(df.columns):
    worksheet.write(0, col, header)

# Write data
for row_idx, row in df.iterrows():
    for col_idx, value in enumerate(row):
        # Handle NaN values
        if pd.isna(value):
            worksheet.write(row_idx + 1, col_idx, '')
        else:
            worksheet.write(row_idx + 1, col_idx, value)

# Debug column headers to ensure we have the right indices
print("\nColumn headers mapping:")
headers_list = list(df.columns)
for col_idx, header in enumerate(headers_list):
    print(f"Column {chr(65 + col_idx)} ({col_idx}): {header}")

# Define formats using xlsxwriter
header_format = workbook.add_format({
    'bold': True,
    'align': 'center',
    'valign': 'vcenter',
    'bg_color': '#D6EAF8',
    'border': 1
})

date_format = workbook.add_format({
    'num_format': 'mm/dd/yyyy',
    'align': 'center'
})

currency_format = workbook.add_format({
    'num_format': '$#,##0.00',
    'align': 'center'
})

currency_format_0 = workbook.add_format({
    'num_format': '$#,##0',
    'align': 'center'
})

number_format = workbook.add_format({
    'num_format': '0',
    'align': 'center'
})

# Apply header formatting and calculate optimal column widths
for col_idx, header in enumerate(headers_list):
    # Set optimal column width based on header length
    optimal_width = max(len(str(header)) + 4, 15)  # At least 15 characters wide
    worksheet.set_column(col_idx, col_idx, optimal_width)
    
    # Apply header formatting
    worksheet.write(0, col_idx, header, header_format)

# Set row height for headers
worksheet.set_row(0, 36)

# Apply column formatting based on column names
column_formats = {}
for col_idx, header in enumerate(headers_list):
    if header == 'Created Date':
        column_formats[col_idx] = date_format
    elif header == 'Carrier Price Per Vehicle':
        column_formats[col_idx] = currency_format
    elif header == 'Vehicle delivery date':
        column_formats[col_idx] = date_format
    elif header == 'Tariff Per Vehicle':
        column_formats[col_idx] = currency_format_0
    elif header == 'Total Carrier Price':
        column_formats[col_idx] = currency_format_0
    elif header == 'Total Price per Mile':
        column_formats[col_idx] = currency_format
    elif header == 'Customer Payment Total Tariff':
        column_formats[col_idx] = currency_format_0
    elif header == 'Accepted by Carrier Date':
        column_formats[col_idx] = date_format
    elif header == 'Carrier Pickup Scheduled At':
        column_formats[col_idx] = date_format
    elif header == 'Pickup ZIP':
        column_formats[col_idx] = number_format
    elif header == 'Delivery Completed At':
        column_formats[col_idx] = date_format
    elif header == 'Adjusted delivery date':
        column_formats[col_idx] = date_format
    elif header == 'Delivery ZIP':
        column_formats[col_idx] = number_format

print("\nColumns to be formatted:")
for col_idx, format_obj in column_formats.items():
    header = headers_list[col_idx]
    print(f"Column {chr(65 + col_idx)} ({col_idx}): {header} - Format applied")

# Apply formatting to data columns
for col_idx, format_obj in column_formats.items():
    # Apply format to the entire column (starting from row 2)
    worksheet.set_column(col_idx, col_idx, None, format_obj)

# Freeze panes at A2 (freeze the header row)
worksheet.freeze_panes(1, 0)

# Enable filters for all columns
last_row = len(df)
last_col = len(df.columns) - 1
worksheet.autofilter(0, 0, last_row, last_col)

# Create pivot table on a new sheet
pivot_sheet = workbook.add_worksheet('Pivot Table')

# Add pivot table headers - Bold, center-aligned
headers = ['Vehicle delivery date', 'VIN #', 'Tariff Per Vehicle', 'Total Carrier Price', 
           'Margin $', 'Margin %', 'Avg Margin Per Unit']

# Set column widths and write headers
for col, header in enumerate(headers):
    pivot_sheet.set_column(col, col, 20)
    pivot_sheet.write(0, col, header, header_format)

# Set header row height
pivot_sheet.set_row(0, 24)

# Ensure numeric columns are properly converted
df['Tariff Per Vehicle'] = pd.to_numeric(df['Tariff Per Vehicle'], errors='coerce').fillna(0)
df['Total Carrier Price'] = pd.to_numeric(df['Total Carrier Price'], errors='coerce').fillna(0)

# Sort the dates before grouping
# First, identify dates in MM/DD/YYYY format for custom sorting
def extract_date_components(date_str):
    if pd.isna(date_str) or date_str == 'Unknown':
        return (9999, 12, 31)  # Put Unknown at the end
    
    if '/' in str(date_str):
        parts = str(date_str).split('/')
        if len(parts) == 3:
            try:
                month = int(parts[0])
                day = int(parts[1])
                year = int(parts[2])
                return (year, month, day)
            except ValueError:
                pass
    
    # Default for non-date strings - sort alphabetically
    return (9999, 12, 30)

# Apply standardization again to ensure all dates have the same format
# This is critical to avoid duplicates in the pivot table
df['Vehicle delivery date'] = df['Vehicle delivery date'].apply(standardize_date)

# Add a sorting key column
df['date_sort_key'] = df['Vehicle delivery date'].apply(lambda x: extract_date_components(x))

# Sort the dataframe
df = df.sort_values('date_sort_key')

# Calculate data for pivot table
# Group by Vehicle delivery date and aggregate values, keeping the original order
pivot_data = df.groupby('Vehicle delivery date', as_index=False, dropna=False, observed=True).agg({
    'VIN #': 'count',
    'Tariff Per Vehicle': 'sum',
    'Total Carrier Price': 'sum'
})

print("\nChecking for duplicate dates:")
all_pivot_dates = pivot_data['Vehicle delivery date'].tolist()
print(f"Unique date formats: {len(set(all_pivot_dates))}")
for i, date in enumerate(all_pivot_dates[:10]):
    print(f"  Date {i+1}: {date} (type: {type(date)})")

# Sort the pivot data by the date components
pivot_data['date_sort_key'] = pivot_data['Vehicle delivery date'].apply(lambda x: extract_date_components(x))
pivot_data = pivot_data.sort_values('date_sort_key')
pivot_data = pivot_data.drop('date_sort_key', axis=1)

# Print pivot data information for debugging
print("\nPivot table - date ranges:")
print(f"First date: {pivot_data['Vehicle delivery date'].iloc[0]}")
print(f"Last date: {pivot_data['Vehicle delivery date'].iloc[-1]}")
print(f"Number of unique dates in pivot: {len(pivot_data)}")

# Verify pivot data contains all expected records
total_records_in_df = len(df)
total_records_in_pivot = pivot_data['VIN #'].sum()
print(f"\nOriginal data has {total_records_in_df} records")
print(f"Pivot table represents {total_records_in_pivot} records")

# Initialize variables to calculate grand totals
grand_total_vin = 0
grand_total_tariff = 0
grand_total_carrier_price = 0
grand_total_margin = 0
grand_total_avg_margin = 0

# Define formats for pivot table
pivot_date_format = workbook.add_format({'num_format': 'mm/dd/yyyy', 'align': 'center'})
pivot_currency_format_0 = workbook.add_format({'num_format': '$#,##0', 'align': 'center'})
pivot_currency_format_2 = workbook.add_format({'num_format': '$#,##0.00', 'align': 'center'})
pivot_percent_format = workbook.add_format({'num_format': '0.0%', 'align': 'center'})
pivot_number_format = workbook.add_format({'align': 'center'})

# Add calculated columns
row_num = 1
for index, row in pivot_data.iterrows():
    # Vehicle delivery date - format as date
    pivot_sheet.write(row_num, 0, row['Vehicle delivery date'], pivot_date_format)
    
    # VIN # count
    vin_count = row['VIN #']
    grand_total_vin += vin_count
    pivot_sheet.write(row_num, 1, vin_count, pivot_number_format)
    
    # Tariff Per Vehicle sum - currency format with 0 decimal places
    tariff = row['Tariff Per Vehicle']
    grand_total_tariff += tariff
    pivot_sheet.write(row_num, 2, tariff, pivot_currency_format_0)
    
    # Total Carrier Price sum - currency format with 0 decimal places
    carrier_price = row['Total Carrier Price']
    grand_total_carrier_price += carrier_price
    pivot_sheet.write(row_num, 3, carrier_price, pivot_currency_format_0)
    
    # Calculate Margin $ (Tariff Per Vehicle - Total Carrier Price) - currency format
    margin_dollars = tariff - carrier_price
    grand_total_margin += margin_dollars
    pivot_sheet.write(row_num, 4, margin_dollars, pivot_currency_format_0)
    
    # Calculate Margin % (Margin $ / Tariff Per Vehicle) - percentage format with 1 decimal place
    margin_percent = margin_dollars / tariff if tariff != 0 else 0
    pivot_sheet.write(row_num, 5, margin_percent, pivot_percent_format)
    
    # Calculate Avg Margin Per Unit (Margin $ / VIN # count) - currency format
    avg_margin_per_unit = margin_dollars / vin_count if vin_count != 0 else 0
    grand_total_avg_margin += (avg_margin_per_unit * vin_count)  # Weighted average for grand total
    pivot_sheet.write(row_num, 6, avg_margin_per_unit, pivot_currency_format_2)
    
    row_num += 1

# Add Grand Total row at the bottom of the pivot table
grand_total_row = row_num
# Calculate overall metrics for the grand total row
overall_margin_percent = grand_total_margin / grand_total_tariff if grand_total_tariff != 0 else 0
overall_avg_margin_per_unit = grand_total_margin / grand_total_vin if grand_total_vin != 0 else 0

# Define grand total format
grand_total_format = workbook.add_format({
    'bold': True,
    'font_size': 14,
    'align': 'center',
    'bg_color': '#D6EAF8',
    'top': 2,
    'bottom': 2
})

grand_total_currency_0 = workbook.add_format({
    'num_format': '$#,##0',
    'bold': True,
    'font_size': 14,
    'align': 'center',
    'bg_color': '#D6EAF8',
    'top': 2,
    'bottom': 2
})

grand_total_currency_2 = workbook.add_format({
    'num_format': '$#,##0.00',
    'bold': True,
    'font_size': 14,
    'align': 'center',
    'bg_color': '#D6EAF8',
    'top': 2,
    'bottom': 2
})

grand_total_percent = workbook.add_format({
    'num_format': '0.0%',
    'bold': True,
    'font_size': 14,
    'align': 'center',
    'bg_color': '#D6EAF8',
    'top': 2,
    'bottom': 2
})

# Add grand total row
pivot_sheet.write(grand_total_row, 0, "Grand Total", grand_total_format)
pivot_sheet.write(grand_total_row, 1, grand_total_vin, grand_total_format)
pivot_sheet.write(grand_total_row, 2, grand_total_tariff, grand_total_currency_0)
pivot_sheet.write(grand_total_row, 3, grand_total_carrier_price, grand_total_currency_0)
pivot_sheet.write(grand_total_row, 4, grand_total_margin, grand_total_currency_0)
pivot_sheet.write(grand_total_row, 5, overall_margin_percent, grand_total_percent)
pivot_sheet.write(grand_total_row, 6, overall_avg_margin_per_unit, grand_total_currency_2)

# Freeze panes on the pivot table headers
pivot_sheet.freeze_panes(1, 0)

# Create a hidden data sheet for chart data
chart_data_sheet = workbook.add_worksheet('_ChartData')
chart_data_sheet.hide()

# Write chart data headers
chart_data_sheet.write(0, 0, 'Vehicle delivery date')
chart_data_sheet.write(0, 5, 'Margin %')

# Write chart data
for idx, (index, row) in enumerate(pivot_data.iterrows(), start=1):
    chart_data_sheet.write(idx, 0, row['Vehicle delivery date'])
    
    # Calculate margin percentage
    tariff = row['Tariff Per Vehicle']
    carrier_price = row['Total Carrier Price']
    margin_dollars = tariff - carrier_price
    margin_percent = margin_dollars / tariff if tariff != 0 else 0
    chart_data_sheet.write(idx, 5, margin_percent)

last_row = len(pivot_data)

# Create Chart Tab with Margin % Trend
chart_sheet = workbook.add_chartsheet('Margin Trend Chart')

# Create the line chart
pivot_chart_margin_pct = workbook.add_chart({'type': 'line'})
pivot_chart_margin_pct.add_series({
    'name': 'Margin %',
    'categories': ['_ChartData', 1, 0, last_row, 0],  # Date range (x-axis)
    'values': ['_ChartData', 1, 5, last_row, 5],  # Margin % values (y-axis)
    'marker': {'type': 'circle', 'size': 5},
    'line': {'width': 2.5, 'color': '#5B9BD5'},  # Light blue
})
pivot_chart_margin_pct.set_title({'name': 'Margin % Trend'})
pivot_chart_margin_pct.set_x_axis({
    'name': 'Delivery Date',
    'num_format': 'mm/dd',
    'num_font': {'rotation': 45},
})
pivot_chart_margin_pct.set_y_axis({
    'name': 'Percentage',
    'num_format': '0.0%',
})
pivot_chart_margin_pct.set_legend({'position': 'top'})
pivot_chart_margin_pct.set_size({'width': 500, 'height': 300})

# Add the chart to the chart sheet
chart_sheet.set_chart(pivot_chart_margin_pct)

# Create Top Cities Sheet
city_sheet = workbook.add_worksheet('Top Cities')

# Create separate dictionaries to track pickup and delivery cities with states
pickup_cities = {}
delivery_cities = {}

# Count occurrences of each city-state combination
for index, row in df.iterrows():
    pickup_city = row.get('Pickup City', '')
    pickup_state = row.get('Pickup State', '')
    delivery_city = row.get('Delivery City', '')
    delivery_state = row.get('Delivery State', '')
    
    # Skip blank or invalid city names
    if isinstance(pickup_city, str) and pickup_city.strip() and isinstance(pickup_state, str) and pickup_state.strip():
        pickup_city = pickup_city.strip()
        pickup_state = pickup_state.strip()
        location_key = f"{pickup_city}, {pickup_state}"
        pickup_cities[location_key] = pickup_cities.get(location_key, 0) + 1
    
    if isinstance(delivery_city, str) and delivery_city.strip() and isinstance(delivery_state, str) and delivery_state.strip():
        delivery_city = delivery_city.strip()
        delivery_state = delivery_state.strip()
        location_key = f"{delivery_city}, {delivery_state}"
        delivery_cities[location_key] = delivery_cities.get(location_key, 0) + 1

# Sort cities by count (most frequent first)
pickup_city_ranking = sorted(pickup_cities.items(), key=lambda x: x[1], reverse=True)
delivery_city_ranking = sorted(delivery_cities.items(), key=lambda x: x[1], reverse=True)

# Add headers for Top Pickup Cities table
pickup_headers = ['Rank', 'Pickup Location', 'Count', 'Percentage']
for col, header in enumerate(pickup_headers, start=1):
    cell = city_sheet.cell(row=1, column=col)
    cell.value = header
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = light_blue_fill
    # Make the city column wider to accommodate city+state format
    if col == 2:  # Location column
        city_sheet.column_dimensions[get_column_letter(col)].width = 25
    else:
        city_sheet.column_dimensions[get_column_letter(col)].width = 18

# Add headers for Top Delivery Cities table
delivery_headers = ['Rank', 'Delivery Location', 'Count', 'Percentage']
for col, header in enumerate(delivery_headers, start=1):
    cell = city_sheet.cell(row=1, column=col + 5)  # Start at column F (5 columns after pickup table)
    cell.value = header
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = light_blue_fill
    # Make the city column wider to accommodate city+state format
    if col == 2:  # Location column
        city_sheet.column_dimensions[get_column_letter(col + 5)].width = 25
    else:
        city_sheet.column_dimensions[get_column_letter(col + 5)].width = 18

# Add "Top Pickup Cities" title
title_cell = city_sheet.cell(row=1, column=1)
title_cell.value = "Top Pickup Locations"
city_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
title_cell.font = Font(bold=True, size=14)
title_cell.alignment = Alignment(horizontal='center', vertical='center')
title_cell.fill = light_blue_fill

# Add "Top Delivery Cities" title
title_cell = city_sheet.cell(row=1, column=6)
title_cell.value = "Top Delivery Locations"
city_sheet.merge_cells(start_row=1, start_column=6, end_row=1, end_column=9)
title_cell.font = Font(bold=True, size=14)
title_cell.alignment = Alignment(horizontal='center', vertical='center')
title_cell.fill = light_blue_fill

# Add actual table headers
for col, header in enumerate(pickup_headers, start=1):
    cell = city_sheet.cell(row=2, column=col)
    cell.value = header
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = light_blue_fill

for col, header in enumerate(delivery_headers, start=1):
    cell = city_sheet.cell(row=2, column=col + 5)
    cell.value = header
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = light_blue_fill

# Calculate total pickups and deliveries for percentage calculation
total_pickups = sum(count for _, count in pickup_city_ranking)
total_deliveries = sum(count for _, count in delivery_city_ranking)

# Add pickup cities data
for idx, (location, count) in enumerate(pickup_city_ranking[:50], start=1):  # Top 50 locations
    row_idx = idx + 2  # Start data from row 3
    
    # Rank
    cell = city_sheet.cell(row=row_idx, column=1)
    cell.value = idx
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(size=12)  # Set font size to 12
    
    # Location (City, State)
    cell = city_sheet.cell(row=row_idx, column=2)
    cell.value = location
    cell.alignment = Alignment(horizontal='left')
    cell.font = Font(size=12)  # Set font size to 12
    
    # Count
    cell = city_sheet.cell(row=row_idx, column=3)
    cell.value = count
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(size=12)  # Set font size to 12
    
    # Percentage
    percentage = (count / total_pickups) * 100 if total_pickups > 0 else 0
    cell = city_sheet.cell(row=row_idx, column=4)
    cell.value = percentage / 100  # Excel will format as percentage
    cell.number_format = '0.0%'
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(size=12)  # Set font size to 12

# Add delivery cities data
for idx, (location, count) in enumerate(delivery_city_ranking[:50], start=1):  # Top 50 locations
    row_idx = idx + 2  # Start data from row 3
    
    # Rank
    cell = city_sheet.cell(row=row_idx, column=6)
    cell.value = idx
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(size=12)  # Set font size to 12
    
    # Location (City, State)
    cell = city_sheet.cell(row=row_idx, column=7)
    cell.value = location
    cell.alignment = Alignment(horizontal='left')
    cell.font = Font(size=12)  # Set font size to 12
    
    # Count
    cell = city_sheet.cell(row=row_idx, column=8)
    cell.value = count
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(size=12)  # Set font size to 12
    
    # Percentage
    percentage = (count / total_deliveries) * 100 if total_deliveries > 0 else 0
    cell = city_sheet.cell(row=row_idx, column=9)
    cell.value = percentage / 100  # Excel will format as percentage
    cell.number_format = '0.0%'
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(size=12)  # Set font size to 12

# Apply consistent column widths to Top Cities tab
city_column_widths = {
    1: 10,  # Rank
    2: 32,  # Pickup Location
    3: 15,  # Count
    4: 15,  # Percentage
    6: 10,  # Rank
    7: 32,  # Delivery Location
    8: 15,  # Count
    9: 15,  # Percentage
}

for col_idx, width in city_column_widths.items():
    city_sheet.column_dimensions[get_column_letter(col_idx)].width = width

# Filtering has been removed from both pickup and delivery tables as requested

# Add light gray alternating row colors
for row_idx in range(3, max(len(pickup_city_ranking), len(delivery_city_ranking)) + 3):
    if row_idx % 2 == 1:  # Odd rows (3, 5, 7, etc.)
        light_gray_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        for col_idx in range(1, 5):  # Pickup cities columns
            cell = city_sheet.cell(row=row_idx, column=col_idx)
            cell.fill = light_gray_fill
        for col_idx in range(6, 10):  # Delivery cities columns
            cell = city_sheet.cell(row=row_idx, column=col_idx)
            cell.fill = light_gray_fill

# Add borders to tables
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# Apply borders to pickup cities table
max_pickup_rows = min(len(pickup_city_ranking) + 2, 52)  # Header + top 50 cities
for row_idx in range(2, max_pickup_rows + 1):
    for col_idx in range(1, 5):
        city_sheet.cell(row=row_idx, column=col_idx).border = thin_border

# Apply borders to delivery cities table
max_delivery_rows = min(len(delivery_city_ranking) + 2, 52)  # Header + top 50 cities
for row_idx in range(2, max_delivery_rows + 1):
    for col_idx in range(6, 10):
        city_sheet.cell(row=row_idx, column=col_idx).border = thin_border

# Freeze the headers
city_sheet.freeze_panes = 'A3'

# Create Carrier Metrics Sheet
carrier_sheet = workbook.create_sheet(title='Carrier Metrics')

# Convert date columns to datetime for delivery time calculation
for date_col in ['Accepted by Carrier Date', 'Delivery Completed At']:
    if date_col in df.columns:
        df[date_col] = pd.to_datetime(df[date_col], errors='coerce')

# Calculate delivery time in days
df['Delivery Time (Days)'] = np.nan
mask = (~df['Accepted by Carrier Date'].isna()) & (~df['Delivery Completed At'].isna())
df.loc[mask, 'Delivery Time (Days)'] = (df.loc[mask, 'Delivery Completed At'] - df.loc[mask, 'Accepted by Carrier Date']).dt.days

# Ensure Carrier Price Per Vehicle is numeric
df['Carrier Price Per Vehicle'] = pd.to_numeric(df['Carrier Price Per Vehicle'], errors='coerce').fillna(0)

# Group by carrier
carrier_metrics = df.groupby('Carrier Name').agg({
    'VIN #': 'count',
    'Carrier Price Per Vehicle': 'mean',
    'Delivery Time (Days)': lambda x: x[x >= 0].mean()  # Only include non-negative delivery times
}).reset_index()

# Sort by VIN count (descending)
carrier_metrics = carrier_metrics.sort_values('VIN #', ascending=False).reset_index(drop=True)

# Add rank column
carrier_metrics.insert(0, 'Rank', range(1, len(carrier_metrics) + 1))

# Add headers to the carrier metrics sheet
headers = ['Rank', 'Carrier Name', 'Total VINs Shipped', 'Avg. Price Per VIN', 'Avg. Delivery Time (Days)']

# Set optimal column widths based on content
optimal_widths = {
    1: 12,  # Rank
    2: 45,  # Carrier Name - wider to accommodate long carrier names
    3: 25,  # Total VINs Shipped
    4: 25,  # Avg. Price Per VIN
    5: 30,  # Avg. Delivery Time (Days)
}

for col, header in enumerate(headers, start=1):
    cell = carrier_sheet.cell(row=1, column=col)
    cell.value = header
    cell.font = Font(bold=True, size=14)  # Set font size to 14 for headers
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.fill = light_blue_fill
    
    # Set column width based on the optimal widths dictionary
    carrier_sheet.column_dimensions[get_column_letter(col)].width = optimal_widths.get(col, 20)

# Increase row height for header to accommodate text wrapping
carrier_sheet.row_dimensions[1].height = 45

# Add data to carrier metrics sheet
for row_idx, row in enumerate(carrier_metrics.iterrows(), start=2):
    index, data = row
    
    # Rank
    rank = row_idx - 1  # Rank starts at 1
    cell = carrier_sheet.cell(row=row_idx, column=1)
    cell.value = rank
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(size=12)  # Set font size to 12
    
    # Carrier Name
    carrier_name = data['Carrier Name']
    cell = carrier_sheet.cell(row=row_idx, column=2)
    cell.value = carrier_name
    cell.alignment = Alignment(horizontal='left')
    cell.font = Font(size=12)  # Set font size to 12
    
    # Total VINs Shipped
    vin_count = data['VIN #']
    cell = carrier_sheet.cell(row=row_idx, column=3)
    cell.value = vin_count
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(size=12)  # Set font size to 12
    
    # Avg. Price Per VIN
    avg_price = data['Carrier Price Per Vehicle']
    cell = carrier_sheet.cell(row=row_idx, column=4)
    cell.value = avg_price
    cell.number_format = '$#,##0.00'
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(size=12)  # Set font size to 12
    
    # Avg. Delivery Time (Days)
    # Handle NaN values
    delivery_time = data['Delivery Time (Days)']
    cell = carrier_sheet.cell(row=row_idx, column=5)
    if pd.isna(delivery_time):
        cell.value = "N/A"
    else:
        cell.value = delivery_time
        cell.number_format = '0.0'
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(size=12)  # Set font size to 12

# Add alternating row colors
for row_idx in range(2, len(carrier_metrics) + 2):
    if row_idx % 2 == 0:  # Even rows
        light_gray_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        for col_idx in range(1, 6):
            carrier_sheet.cell(row=row_idx, column=col_idx).fill = light_gray_fill

# Add borders to table
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# Apply borders to entire table
for row_idx in range(1, len(carrier_metrics) + 2):
    for col_idx in range(1, 6):
        carrier_sheet.cell(row=row_idx, column=col_idx).border = thin_border

# Freeze the header row
carrier_sheet.freeze_panes = 'A2'

# Add filter to the header row
max_row = len(carrier_metrics) + 1  # +1 for the header row
carrier_sheet.auto_filter.ref = f"A1:E{max_row}"

# Close the workbook
workbook.close()

print("\nExcel file 'sales_report.xlsx' created successfully with all requested formatting, calculations, and chart.") 